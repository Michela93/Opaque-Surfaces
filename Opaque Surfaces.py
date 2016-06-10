# -*- coding: utf-8 -*-
from openpyxl import * # we need openpyxl to connect Phyton with Excel file
import os
"""This Python script is going to allow the User to calculate the U_value, the cooling factor (CF), the heating factor (HF), the cooling load (Q_cooling) and the heating 
load (Q_heating)  of the main Opaque Surfaces of a building: ceiling, walls, floor, doors. These results will be summarized in a new excel file called OpaqueSurfaceCharacteristics 
which will be automatically created as the script proceeds with the calculations.
This script is somehow interactive and will need some inputs  from the user (both in the excel file and during the running procedure of the program).
First of all the user will have to open the excel  file “DataOS” and fill in some of the cells depending on the characteristics of the building components:
_ In the sheet HouseCharacteristics, the user will have to insert the area of each surface and the DR value (K) which is the cooling daily range for the building.
_ In the sheet Wall_Input,  the user will have to insert both the ID name of each material that is part of the wall’s layers (such IDs can be found in the R_values sheet) and the thickness of each layer (mm).
_ In the sheet Ceiling_Input,  the user will have to insert both the ID name of each material that is part of the ceiling’s layers (such IDs can be found in the R_values sheet), the thickness of each layer (mm) 
and the color (white, light, medium, dark).
_In the sheet Floor_Input,  the user will have to insert both the ID name of each material that is part of the floor’s layers (such IDs can be found in the R_values sheet) and the thickness of each layer (mm).
_In the sheet Door_Input,  the user will have to insert both the ID name of each material that is part of the door’s layers (such IDs can be found in the R_values sheet) and the thickness of each layer (mm).
Once these cells are filled out, the user can run the program. There will be the creation of the new excel file OpaqueSurfaceCharacteristics.
Now the python interface will interact whit the user by asking at first if the building’s characteristics have to be considered in either summer of winter conditions, then other  4 questions regarding some specific 
characteristics of each surface: this will allow the calculation of the CF.
Once such questions are answered, the user will obtain all the results in terms of U_value, HF, CF, Q_heating and Q_cooling , all summarized in the HouseCharacteristics sheet found in the OpaqueSurfaceCharacteristics file."""

os.chdir("C:\Users\Michela\Searches\Magistrale\Energy and environmental technologies for building systems\Opaque Surface") # here is defined our path to Excel file
ExcelFile = load_workbook("DataOS.xlsx") # name of the excel file

Season=raw_input("Is it summer or winter? summer=1 winter=2 : ")
if Season=="1":
    R_out=0.044
else:
    R_out=0.03

R_in=0.12

##################################WALL##################################
Wall_input = ExcelFile.get_sheet_by_name("Wall_input") # name of the excel sheet from which we extract the useful values for the calculations
name_cells = Wall_input.columns[0][1:] # for the calculation we need to take values from all the rows in first column

for cell in name_cells: # with this "for" loop we are printing names of layers given by the user
    name = cell.value
    print name
layers = [] 

Number_of_layers=len(name_cells) # number of layers considered
layers=list()  #we create an empty list of the layers of the wall 
for index in range(1,Number_of_layers+1):   # with this "for" loop, in the range given by the number of the layers specified in Excel, we fullfil the new list 
    layer = (Wall_input.columns[0][index].value).encode('utf-8')  #the value contained in the cell is a word
    thickness = float(Wall_input.columns[1][index].value)  #the value contained in the cell is a number
    wall_layer = {"layer":layer,"thickness":thickness}  
    layers.append(wall_layer)  #list is filled up with the names of the layers and thicknesses of the layers  

#R_values
ExcelFile = load_workbook("DataOS.xlsx")  #we are using this excel file 
R_values = ExcelFile.get_sheet_by_name("R_values")   # now we change excel sheet to extract R_values

layer_ID =R_values.columns[1][8:] #we call the different layers ID , starting from the row number 9 in the first column because the previous values were Rin and Rout values, together with air gaps values
layer_materials = list() #then we create a list with the materials from tables for our calculation
for cell in layer_ID: # in this loop we fullfil the new empty list with names of the layers´ materials
    material = (cell.value).encode('utf-8')  
    layer_materials.append(material) 
    
R_tot_wall= 0  # each material of our layers has different R value, given by table. Value of R depends also from the thickness of the layer. We have to define the right R values for next calculation of the U factor.
for layer in layers:   # in this loop we specify R value of air gap between the layers, due to it´s thickness.
    print layer
    layer_material=layer["layer"]
    layer_thickness=layer["thickness"]
    if (layer_material=="air_space_13") and (layer_thickness==13):  #if thickness of the air layer is "13", then R value is given by tables. 
        R_wall=R_values[4][3].value  # we extract the R value
        layer["R_wall"]=R_wall
    elif (layer_material=="air_space_20") and (layer_thickness==20):
        R_wall=R_values.rows[5][3].value
        layer["R_wall"]=R_wall
    elif (layer_material=="air_space_40") and (layer_thickness==40):
        R_wall=R_values.rows[6][3].value
        layer["R_wall"]=R_wall
    elif (layer_material=="air_space_90") and (layer_thickness==90):
        R_wall=R_values.rows[7][3].value 
        layer["R_wall"]=R_wall
    else:  # from here on, we consider all other materials for layers 
        row=layer_materials.index(layer_material)+8 
        R_Standard=R_values.rows[row][3].value
        L_standard=R_values.rows[row][2].value
        R_wall =R_Standard*layer_thickness/L_standard 
        layer["R_wall"]=R_wall

    R_tot_wall += R_wall  # for calculation of the U factor of the whole wall we consider only R total
U_wall=1/(R_tot_wall+R_out+R_in)  # calculation of the U factor 
print U_wall  #here we print our result = U factor of the wall
print layers  # we print number of the layers in the wall


##################################CEILING##################################
# for specification of the R values of each type of ceiling layer, we use the same manner like in the previous wall section. However for the layers of the ceiling we have to also consider alpha, which is also given by tables.
 
Ceiling_input = ExcelFile.get_sheet_by_name("Ceiling_input") # for ceiling inputs we have to change the sheet of Excel 
name_cells = Ceiling_input.columns[0][1:] # we call all layers in our ceiling given by user

for cell in name_cells:  
    name = cell.value
layers = []

Number_of_layers=len(name_cells)
layers=list()
for index in range(1,Number_of_layers+1):
    layer = (Ceiling_input.columns[0][index].value).encode('utf-8')
    thickness = float(Ceiling_input.columns[1][index].value)
    colour = (Ceiling_input.columns[2][index].value)
    ceiling_layer = {"layer":layer,"thickness":thickness, "colour":colour}
    layers.append(ceiling_layer)   #we create the dictionary of layers 


#R_values
ExcelFile = load_workbook("DataOS.xlsx")
R_values = ExcelFile.get_sheet_by_name("R_values")   

layer_ID =R_values.columns[1][8:]
layer_materials = list()

for cell in layer_ID:
    material = (cell.value).encode('utf-8')
    layer_materials.append(material)
R_tot_ceiling= 0 
for layer in layers:
    print layer
    layer_material=layer["layer"]
    layer_thickness=layer["thickness"]
    if (layer_material=="air_space_13") and (layer_thickness==13):
        R_ceiling=R_values[4][3].value
        layer["R_ceiling"]=R_ceiling
    elif (layer_material=="air_space_20") and (layer_thickness==20):
        R_ceiling=R_values.rows[5][3].value
        layer["R_ceiling"]=R_ceiling
    elif (layer_material=="air_space_40") and (layer_thickness==40):
        R_ceiling=R_values.rows[6][3].value
        layer["R_ceiling"]=R_ceiling
    elif (layer_material=="air_space_90") and (layer_thickness==90):
        R_ceiling=R_values.rows[7][3].value 
        layer["R_ceiling"]=R_ceiling
    else:
        row=layer_materials.index(layer_material)+8
        R_Standard=R_values.rows[row][3].value
        L_standard=R_values.rows[row][2].value
        R_ceiling =R_Standard*layer_thickness/L_standard 
        layer["R_ceiling"]=R_ceiling
        
    
    R_tot_ceiling += R_ceiling
U_ceiling=1/(R_tot_ceiling+R_out+R_in)
print U_ceiling

# For ceiling materials have to be defined also the alpha values. These values are picked from the table "Alpha_values" in the Excel.
ExcelFile = load_workbook("DataOS.xlsx") # the file where we can find the alpha values
Alpha_values = ExcelFile.get_sheet_by_name("Alpha_values") # we call sheet with the alpha values 

layer_ID =Alpha_values.columns[1][8:]
layer_materials = list() # we create an empty list of the materials of layers used in the ceiling 

for cell in layer_ID: # the for-loop fullfils the list 
    material = (cell.value).encode('utf-8')
    layer_materials.append(material)

for layer in layers: # in this loop we specify alpha values depending on the ceiling color
    layer_material=layer["layer"]
    layer_colour=layer["colour"]
    if (layer_colour=="white") :
        alpha=Alpha_values.rows[1][2].value
        layer["alpha"]=alpha
    elif (layer_colour=="light") :
        alpha=Alpha_values.rows[1][3].value
        layer["alpha"]=alpha
    elif (layer_colour=="medium") :
        alpha=Alpha_values.rows[1][4].value
        layer["alpha"]=alpha    
    elif (layer_colour=="dark") :
        alpha=Alpha_values.rows[1][5].value
        layer["alpha"]=alpha        

print alpha
print layers


##################################FLOOR##################################
Floor_input = ExcelFile.get_sheet_by_name("Floor_input")  # for specification of the R values of each type of floor layer, we use the same manner like in the previous sections 
name_cells = Floor_input.columns[0][1:]

for cell in name_cells:
    name = cell.value
    print name
layers = []

Number_of_layers=len(name_cells)
layers=list()
for index in range(1,Number_of_layers+1):
    layer = (Floor_input.columns[0][index].value).encode('utf-8')
    thickness = float(Floor_input.columns[1][index].value)
    floor_layer = {"layer":layer,"thickness":thickness}
    layers.append(floor_layer) 

#R_values
ExcelFile = load_workbook("DataOS.xlsx")
R_values = ExcelFile.get_sheet_by_name("R_values")   

layer_ID =R_values.columns[1][8:]
layer_materials = list()

for cell in layer_ID:
    material = (cell.value).encode('utf-8')
    layer_materials.append(material)
R_tot_floor= 0 
for layer in layers:
    print layer
    layer_material=layer["layer"]
    layer_thickness=layer["thickness"]
    if (layer_material=="air_space_13") and (layer_thickness==13):
        R_floor=R_values[4][3].value
        layer["R_floor"]=R_floor
    elif (layer_material=="air_space_20") and (layer_thickness==20):
        R_floor=R_values.rows[5][3].value
        layer["R_floor"]=R_floor
    elif (layer_material=="air_space_40") and (layer_thickness==40):
        R_floor=R_values.rows[6][3].value
        layer["R_floor"]=R_floor
    elif (layer_material=="air_space_90") and (layer_thickness==90):
        R_floor=R_values.rows[7][3].value 
        layer["R_floor"]=R_floor
    else:
        row=layer_materials.index(layer_material)+8
        R_Standard=R_values.rows[row][3].value
        L_standard=R_values.rows[row][2].value
        R_floor =R_Standard*layer_thickness/L_standard 
        layer["R_floor"]=R_floor

    R_tot_floor += R_floor
U_floor=1/R_tot_floor
print U_floor
print layers

##################################DOOR##################################
Door_input = ExcelFile.get_sheet_by_name("Door_input") # for specification of the R values of each type of door layer, we use the same manner like in the previous sections 
name_cells = Door_input.columns[0][1:]

for cell in name_cells:
    name = cell.value
    print name
layers = []

Number_of_layers=len(name_cells)
layers=list()
for index in range(1,Number_of_layers+1):
    layer = (Door_input.columns[0][index].value).encode('utf-8')
    thickness = float(Door_input.columns[1][index].value)
    door_layer = {"layer":layer,"thickness":thickness}
    layers.append(door_layer)   

#R_values
ExcelFile = load_workbook("DataOS.xlsx")
R_values = ExcelFile.get_sheet_by_name("R_values")   

layer_ID =R_values.columns[1][8:]
layer_materials = list()

for cell in layer_ID:
    material = (cell.value).encode('utf-8')
    layer_materials.append(material)
R_tot_door= 0 
for layer in layers:
    print layer
    layer_material=layer["layer"]
    layer_thickness=layer["thickness"]
    if (layer_material=="air_space_13") and (layer_thickness==13):
        R_door=R_values[4][3].value
        layer["R_door"]=R_door
    elif (layer_material=="air_space_20") and (layer_thickness==20):
        R_door=R_values.rows[5][3].value
        layer["R_door"]=R_door
    elif (layer_material=="air_space_40") and (layer_thickness==40):
        R_door=R_values.rows[6][3].value
        layer["R_door"]=R_door
    elif (layer_material=="air_space_90") and (layer_thickness==90):
        R_door=R_values.rows[7][3].value 
        layer["R_door"]=R_door
    else:
        row=layer_materials.index(layer_material)+8
        R_Standard=R_values.rows[row][3].value
        L_standard=R_values.rows[row][2].value
        R_door =R_Standard*layer_thickness/L_standard 
        layer["R_door"]=R_door

    R_tot_door += R_door
U_door=1/(R_tot_door+R_out+R_in)
print U_door
print layers



#Thanks to the following command we are able to insert the U_values of in a new Excel file called OpaqueSurfaceCharacteristic in the sheet HouseCharacteristics
Results=ExcelFile.get_sheet_by_name("HouseCharacteristics")
Results.columns[5][2].value=U_wall
Results.columns[5][1].value=U_ceiling
Results.columns[6][1].value=alpha
Results.columns[5][4].value=U_floor
Results.columns[5][3].value=U_door
ExcelFile.save("OpaqueSurfaceCharacteristic.xlsx")

#################################################################################
##################################HeatingFactor##################################
T_set_heating = 20 # here the user has to insert the values of the given temperatures 
T_set_cooling = 24
T_out_heating = -3
T_out_cooling = 33

deltaT_Heating = T_set_heating-T_out_heating # we calculate the difference between outside temperature and inside temperature for summer and winter
deltaT_cooling = T_out_cooling-T_set_cooling

ExcelFile = load_workbook("OpaqueSurfaceCharacteristic.xlsx") # we use this excel file
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics") # calling sheet
U_cells = HouseCharacteristics.columns[5][1:] # we call the U values in column F for each component
U_values=list() # we create and complete the list with this values by next loop
for cell in U_cells:
    cell_U= (cell.value)
    U_values.append(cell_U)

HF_values=[] # then we have to evaluate the heating factor
for U_n in U_values : #for calculation of heating factor we use the previous list of U values
    HF=U_n*deltaT_Heating # calculation of heating factor
    HF_values.append(HF)  #we obtain a list of results
       
Results=ExcelFile.get_sheet_by_name("HouseCharacteristics") # we put our results of heating factor into Ecxel in sheet:"HouseCharacteristics"
for index in range(0,4): #we fulfill only rows in range
    Results.columns[9][index+1].value=HF_values[index] # we specify the location of the column where we want to save our HF values
    ExcelFile.save("OpaqueSurfaceCharacteristic.xlsx") # saving HF values into the Excel file    

                    
##################################Q_heating##################################    
ExcelFile = load_workbook("OpaqueSurfaceCharacteristic.xlsx") # we consider the new Excel file for next calculations
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics") # specifying the sheet in the Excel file

Q_heating=0
columns_area = HouseCharacteristics.columns[1][1:] # for calculation we need values of component´s areas and we have to specify their location 
area_values=list() #list of our values, which we will use in next calculation
for cell in columns_area:  # fulfilling the list of area values
    area = float(cell.value)
    print area
    area_values.append(area)

columns_HF = HouseCharacteristics.columns[9][1:] # for calculation we need our values of HF, which were calculated by us in previous section. Also we have to define their location to pick them from sheet.
HF_values=list() #list of our values, which we will use in next calculation 
for cell in columns_HF: # fulfilling the list of HF values 
    HF = float(cell.value)
    print HF
    HF_values.append(HF)

Q_heating_values=list() # we create the list with our results of calculation of Opaque surface heating load for each component
for i in range(0,4):   # calculating only in the range of ours components 
    Q_heating += (area_values[i]*HF_values[i]) #calculation of Opaque surface heating load
    Q_heating_values.append(Q_heating) #adding values to the list 

Results=ExcelFile.get_sheet_by_name("HouseCharacteristics") # our results of Opaque surface heating load are put into Ecxel in sheet:"HouseCharacteristics"
for index in range(0,4): #we fulfill only rows in range
    Results.columns[10][index+1].value=Q_heating_values[index]   # we specify the location: column where we want to save our Opaque surface heating load values
    ExcelFile.save("OpaqueSurfaceCharacteristic.xlsx")    # saving Opaque surface heating load values into the Excel file    

 
#################################################################################
##################################CoolingFactor##################################
ExcelFile = load_workbook("OpaqueSurfaceCharacteristic.xlsx") #  we consider the Excel file for next calculations

CeilingType=raw_input("Is your ceiling adjacent to vented attic? yes=1 no=2 : ")   # we ask user to specify his ceiling for next calculation 
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics")    #open sheet in the excel file 
alpha_extracted =HouseCharacteristics.columns[6][1].value     # we extract the alpha value we need for calculation
DR=HouseCharacteristics.columns[7][1].value     #we extract value of cooling daily range - DR
U=HouseCharacteristics.columns[5][1].value    #also we extract U value from column 5
if CeilingType=="1":     # due to answer of user to our question we do next calculation
    OF_values = ExcelFile.get_sheet_by_name("OF_values")   #we will use values from this sheet, so we have to call it
    OFt_value= OF_values.columns[2][1].value 
    OFr_value= OF_values.columns[3][1].value 
    OFb_value = (14.3*alpha_extracted-4.5) #OFb value is not given so we have to calculate it, due to given relation 
    CF_ceiling=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR)) # calculation of cooling factor for ceiling adjacent to vented attic
else:      #if user´s answer is diffrent from "1", then we use this calculation, with different values, but in the same sheet
    OF_values = ExcelFile.get_sheet_by_name("OF_values") 
    OFt_value= OF_values.columns[2][1].value 
    OFr_value= OF_values.columns[3][1].value 
    OFb_value = (38.3*alpha_extracted-7)   #OFb calculation for ceiling/roof assembly. Value of OFb is not given so we have to calculate it, due to given relation 
    CF_ceiling=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))  # calculation of cooling factor for ceiling/roof assembly

 # Also for cooling factor of walls there can be different inputs depending on each kind of wall, so we ask the user which type of wall he has and due to this, we again calculate the cooling factor for the wall.
WallType=raw_input("Is your wall exposed to solar radiation or is it shaded? exposed=1 shaded=2 : ") 
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics") 
DR=HouseCharacteristics.columns[7][1].value
U=HouseCharacteristics.columns[5][2].value
if WallType=="1":
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][1].value
    OFr_value= OF_values.columns[3][1].value 
    OFb_value = 8.2 #OFb is given by tables
    CF_wall=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))
else:
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][3].value 
    OFr_value= OF_values.columns[3][3].value 
    OFb_value = 0 #OFb is given by tables 
    CF_wall=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))

# we calculate the Cooling factor of the floor as before   
FloorType=raw_input("Is your floor positioned over an ambient or over a crawlspace? ambient=1 crawlspace=2 : ")
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics") 
DR=HouseCharacteristics.columns[7][1].value
U=HouseCharacteristics.columns[5][4].value
if FloorType=="1":
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][5].value 
    OFr_value= OF_values.columns[3][5].value  
    OFb_value = 0 #OFb is given by tables 
    CF_floor=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))
else:
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][6].value 
    OFr_value= OF_values.columns[3][6].value  
    OFb_value = 0 #OFb is given by tables 
    CF_floor=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))
    

# we calculate the Cooling factor of the door as before 
DoorType=raw_input("Is your door exposed to solar radiation or is it shaded? exposed=1 shaded=2 : ")
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics") 
DR=HouseCharacteristics.columns[7][1].value
U=HouseCharacteristics.columns[5][3].value
if DoorType=="1":
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][1].value 
    OFr_value= OF_values.columns[3][1].value 
    OFb_value = 8.2 #OFb is given by tables 
    CF_door=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))
else:
    OF_values = ExcelFile.get_sheet_by_name("OF_values")
    OFt_value= OF_values.columns[2][3].value 
    OFr_value= OF_values.columns[3][3].value 
    OFb_value = 0 #OFb is given by tables 
    CF_door=(U*(OFt_value*deltaT_cooling+OFb_value+OFr_value*DR))
    
# here we print the cooling factors' results for each component 
print("\n  The CF value for the ceiling is: "+ str(CF_ceiling)+ "\n") 
print("\n  The CF value for the walls is: "+ str(CF_wall)+ "\n") 
print("\n  The CF value for the floor is: "+ str(CF_floor)+ "\n") 
print("\n  The CF value for the door is: "+ str(CF_door)+ "\n") 

Results=ExcelFile.get_sheet_by_name("HouseCharacteristics") # the results of cooling factor are put into Ecxel in sheet:"HouseCharacteristics"
Results.columns[11][1].value=CF_ceiling # we specify the location: column where we want to save our CF values
Results.columns[11][2].value=CF_wall
Results.columns[11][4].value=CF_floor
Results.columns[11][3].value=CF_door
ExcelFile.save("OpaqueSurfaceCharacteristic.xlsx") # saving CF values into the Excel file    

##################################Q_cooling##################################    
ExcelFile = load_workbook("OpaqueSurfaceCharacteristic.xlsx") 
HouseCharacteristics = ExcelFile.get_sheet_by_name("HouseCharacteristics")

Q_cooling=0 
columns_area = HouseCharacteristics.columns[1][1:] # taking values of component´s areas from the sheet
area_values=list() # we create a empty list of values of component´s areas, which we´ll fullfil with next "for" loop
for cell in columns_area:
    area = float(cell.value)
    print area
    area_values.append(area)

columns_CF = HouseCharacteristics.columns[11][1:] # for calculation we also need cooling factors, which we have to extract from the sheet. We create a list of these values, like in the previous case.
CF_values=list()
for cell in columns_CF:
    CF = float(cell.value)
    print CF
    CF_values.append(CF)

Q_cooling_values=list() # we create a list of the calculation of opaque surface cooling loads, which are calculated in the next "for" loop
for i in range(0,4): # we are using values only in the range   
    Q_cooling += (area_values[i]*CF_values[i]) # relation for calculation of opaque surface cooling load for each component
    Q_cooling_values.append(Q_heating) # adding values of our results to the list

Results=ExcelFile.get_sheet_by_name("HouseCharacteristics")# results of Opaque surface cooling load are put into Ecxel in sheet:"HouseCharacteristics"
for index in range(0,4): #we fulfill only rows in range
    Results.columns[12][index+1].value=Q_cooling_values[index] # we specify the location: column where we want to save our Q cooling values
    ExcelFile.save("OpaqueSurfaceCharacteristic.xlsx") # saving Opaque surface cooling load values into the Excel file       






